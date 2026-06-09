from flask import render_template, jsonify
from sqlalchemy import func
from model import db, ユーザ, 所属, 拾得物, 拾得物分類, 拾得物管理状況
from datetime import datetime
import os,re
import openpyxl

ALLOWED_IMAGE_EXT = [".jpg", ".jpeg", ".png", ".gif"]
MAX_IMAGE_SIZE = 2 * 1024 * 1024

def is_empty(value):
    return value is None or value.strip() == ""

def has_dangerous_chars(value):
    if value is None:
        return False
    return bool(re.search(r"[;'\"\\]", value))

def check_required(data, keys):
    errors = []
    for key in keys:
        if is_empty(data.get(key)):
            errors.append(f"{key}は必須です")
    return errors

def check_text_values(data):
    errors = []
    for key, value in data.items():
        if has_dangerous_chars(value):
            errors.append(f"{key}に使用できない記号が含まれています")
    return errors

def check_image(file):
    errors = []
    if file is None or file.filename == "":
        errors.append("画像ファイルを選択してください")
        return errors
    name, ext = os.path.splitext(file.filename)
    if ext.lower() not in ALLOWED_IMAGE_EXT:
        errors.append("画像は jpg, jpeg, png, gif のみ使用できます")
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > MAX_IMAGE_SIZE:
        errors.append("画像サイズは2MB以下にしてください")
    return errors

def top():
    return render_template("top.html")

def get_json(data):
    if data == "user":
        return jsonify(get_user_list())
    elif data == "item":
        return jsonify(get_item_list())
    elif data == "category":
        return jsonify(get_category_list())
    elif data == "dept":
        return jsonify(get_dept_list())

def get_dept_list():
    data = db.session.query(所属).all()
    obj = []
    for i in data:
        obj.append({
            "所属ID":i.ID,
            "所属分類":i.所属分類,
            "所属名":i.所属名, 
            })
    return(obj)

def get_user_list():
    data = db.session.query(ユーザ,所属).filter(ユーザ.所属ID==所属.ID).all()
    obj = []
    for i,j in data:
        obj.append({
            "ID":i.ID,
            "氏名":i.氏名,
            "電話番号":i.電話番号,
            "メールアドレス":i.メールアドレス,
            "住所":i.住所,
            "所属分類":j.所属分類,
            "所属名":j.所属名, 
            })
    return(obj)

def get_item_list(key=""):
    lastdt = db.session.query(
        拾得物管理状況.拾得物ID.label("oid"),
        func.max(拾得物管理状況.変更日時).label("last")
    ).group_by(
        拾得物管理状況.拾得物ID
    ).subquery()
    tmp = db.session.query(
        拾得物管理状況, 
        拾得物, 
        拾得物分類, 
        ユーザ, 
        所属
    ).filter(
        拾得物管理状況.拾得物ID==拾得物.ID
    ).filter(
        拾得物.拾得物分類ID==拾得物分類.ID
    ).filter(
        拾得物管理状況.ユーザID==ユーザ.ID
    ).filter(
        ユーザ.所属ID==所属.ID
    ).filter(
        拾得物管理状況.変更日時==lastdt.c.last
    ).filter(
        拾得物管理状況.拾得物ID==lastdt.c.oid
    )
    if key == "":
        data = tmp.all()
    else:
        data = tmp.filter(拾得物分類.物品名.contains(key)).all()
    obj = []
    for i,j,k,l,m in data:
        obj.append({
            "大分類":k.大分類,
            "物品名":k.物品名,
            "貴重品":k.貴重品,
            "拾得物ID":j.ID,
            "拾得場所":j.拾得場所,
            "色":j.色,
            "特徴":j.特徴,
            "画像":j.画像,
            "拾得物管理状況ID":i.ID,
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            "氏名":l.氏名,
            "所属名":m.所属名, 
            })
    return(obj)

def get_category_list():
    data = db.session.query(拾得物分類).all()
    obj = []
    for i in data:
        obj.append({
            "ID":i.ID,
            "大分類":i.大分類,
            "物品名":i.物品名,
            "頭１":i.頭１,
            "頭２":i.頭２,
            "五十音":i.五十音,
            "貴重品":i.貴重品, 
            })
    return(obj)

def search_item(k):
    j = get_item_list(k)
    return render_template("item_list.html", j=j, n=len(j))

def get_item_detail(id, errors=[]):
    data = db.session.query(拾得物管理状況,ユーザ).filter(拾得物管理状況.ユーザID==ユーザ.ID).filter(拾得物管理状況.拾得物ID==id)
    list = []
    for i,j in data:
        list.append({
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            "氏名":j.氏名,
        })
    result = db.session.query(拾得物,拾得物分類).filter(拾得物.拾得物分類ID==拾得物分類.ID).filter(拾得物.ID==id).first()
    if result is None:
        return "指定された拾得物は存在しません"
    i,j = result
    if i.画像 is None:
        i.画像 = ""
    d = {
        "大分類":j.大分類,
        "物品名":j.物品名,
        "拾得場所":i.拾得場所,
        "色":i.色,
        "特徴":i.特徴,
        "画像":i.画像, 
        "変更履歴":list, 
        "id":id
        }
    return render_template(
        "item_detail.html", 
        d=d, 
        u=get_user_list(), 
        errors=errors
    )

def user_form(errors=[]):
    return render_template(
        "reg_user.html", 
        j=get_dept_list(),
        errors=errors
    )

def user_reg(d):
    tmp = ユーザ(
        氏名 = d["氏名"],
        所属ID = d["所属ID"],
        電話番号 = d["電話番号"],
        メールアドレス = d["メールアドレス"],
        住所= "",
    )
    db.session.add(tmp)
    db.session.commit()

def item_form(errors=[]):
    return render_template(
        "reg_item.html", 
        j=get_category_list(), 
        j2=get_user_list(),
        errors=errors
    )

def item_reg(d, f):
    tmp = 拾得物(
        拾得物分類ID = d["拾得物分類ID"],
        拾得場所 = d["拾得場所"],
        色 = d["色"],
        特徴 = d["特徴"],
        画像 = "",
    )
    db.session.add(tmp)
    db.session.flush()
    n,ext = os.path.splitext(f.filename)
    fn = f"{tmp.ID}{ext}"
    tmp.画像 = fn
    f.save(os.path.join("./static/img/", fn))
    tmp2 = 拾得物管理状況(
        ユーザID = d["ユーザID"],
        拾得物ID = tmp.ID,
        変更日時 = datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
        変更内容 = "拾得",
    )
    db.session.add(tmp2)
    db.session.commit()

def act_reg(d):
    tmp = 拾得物管理状況(
        ユーザID = d["ユーザID"],
        拾得物ID = d["拾得物ID"],
        変更日時 = datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
        変更内容 = d["変更内容"],
    )
    db.session.add(tmp)
    db.session.commit()

def req_item():
    return render_template("req_item.html")

def req_list():
    return render_template("req_list.html")

def dl():
    return render_template("dl.html")

def dl_data():
    lastdt = db.session.query(拾得物管理状況.拾得物ID.label("oid"),func.max(拾得物管理状況.変更日時).label("last")).group_by(拾得物管理状況.拾得物ID).subquery()
    tmp = db.session.query(拾得物管理状況,拾得物,拾得物分類).filter(拾得物管理状況.拾得物ID==拾得物.ID).filter(拾得物.拾得物分類ID==拾得物分類.ID).filter(拾得物管理状況.変更日時==lastdt.c.last).filter(拾得物管理状況.拾得物ID==lastdt.c.oid).filter(拾得物管理状況.変更内容=='拾得').all()
    obj = []
    for i,j,k in tmp:
        obj.append({
            "物品名":k.物品名,
            "拾得物ID":j.ID,
            "拾得場所":j.拾得場所,
            "色":j.色,
            "特徴":j.特徴,
            "拾得物管理状況ID":i.ID,
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            })
    path = os.path.dirname(__file__)
    wb = openpyxl.load_workbook(f"{path}\style05.xlsx")
    pat = ["0","1","2","3","4","5","6","7","8","9","０","１","２","３","４","５","６","７","８","９"]
    i,p = 0,0
    for row in obj:
        if (i + 10) > 46 or i == 0:
            i,p = 0,p+1
            ws = wb.copy_worksheet(wb["template"])
            ws.title = f"拾得物リスト{p}"
        ws.cell(10+i,9).value = row["物品名"]
        if row["物品名"] == "現金":
            j = 0
            for s in row["特徴"][::-1]:
                if not(s in pat): continue
                ws.cell(10+i,8-j).value = s
                print(s)
                j += 1
            ws.cell(10+i,8-j).value = "¥"
        else:
            ws.cell(10+i,11).value = row["特徴"] + row["色"]
        ws.cell(10+i,17).value = row["変更日時"].strftime('%m/%d %H:%M')
        ws.cell(10+i,19).value = row["拾得場所"]
        i += 4
    ws = wb.remove(wb["template"])
    wb.save(f"{path}\out.xlsx")
    return